"""
Build a single .xlsx spreadsheet with all assumptions editable at the top,
and Via Labs + RealReal Genuine + Combined P&L below — all formula-driven.
Upload to Google Drive → Open with Google Sheets and everything works.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ViaLabs_RRG_PL.xlsx")

wb = openpyxl.Workbook()

# ─── Styles ───────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
NORMAL_FONT = Font(name="Arial", size=10)
EDITABLE_FONT = Font(name="Arial", size=10, bold=True, color="1A237E")
RESULT_FONT = Font(name="Arial", size=10, bold=True)

TITLE_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
SECTION_FILL = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
HEADER_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
EDITABLE_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Yellow
TOTAL_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Green
LOSS_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Red tint

THIN_BORDER = Border(
    bottom=Side(style="thin", color="BDBDBD"),
)
THICK_BORDER = Border(
    top=Side(style="medium", color="1A237E"),
    bottom=Side(style="medium", color="1A237E"),
)

YEARS = [2026, 2027, 2028, 2029, 2030]
YEAR_COLS = {2026: "C", 2027: "D", 2028: "E", 2029: "F", 2030: "G"}


def style_section_header(ws, row, text):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left")
    for col in "BCDEFG":
        ws[f"{col}{row}"].fill = SECTION_FILL


def style_title(ws, row, text):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left")
    for col in "BCDEFG":
        ws[f"{col}{row}"].fill = TITLE_FILL


def write_year_headers(ws, row):
    ws[f"A{row}"].value = ""
    ws[f"A{row}"].fill = HEADER_FILL
    ws[f"B{row}"].value = ""
    ws[f"B{row}"].fill = HEADER_FILL
    for i, yr in enumerate(YEARS):
        col = get_column_letter(3 + i)
        c = ws[f"{col}{row}"]
        c.value = yr
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def write_label_row(ws, row, label, values=None, fmt="number", font=None, fill=None, indent=0):
    prefix = "  " * indent
    c = ws[f"A{row}"]
    c.value = prefix + label
    c.font = font or NORMAL_FONT
    if fill:
        for col in "ABCDEFG":
            ws[f"{col}{row}"].fill = fill
    if values:
        for i, val in enumerate(values):
            col = get_column_letter(3 + i)
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.font = font or NORMAL_FONT
            cell.alignment = Alignment(horizontal="center")
            if fmt == "currency":
                cell.number_format = '#,##0'
            elif fmt == "percent":
                cell.number_format = '0.0%'
            elif fmt == "number":
                cell.number_format = '#,##0'
            if fill:
                cell.fill = fill


def write_editable_row(ws, row, label, values, fmt="number", note=""):
    c = ws[f"A{row}"]
    c.value = label
    c.font = EDITABLE_FONT
    ws[f"B{row}"].value = note
    ws[f"B{row}"].font = Font(name="Arial", size=8, italic=True, color="757575")
    for i, val in enumerate(values):
        col = get_column_letter(3 + i)
        cell = ws[f"{col}{row}"]
        cell.value = val
        cell.font = EDITABLE_FONT
        cell.fill = EDITABLE_FILL
        cell.alignment = Alignment(horizontal="center")
        if fmt == "currency":
            cell.number_format = '#,##0'
        elif fmt == "percent":
            cell.number_format = '0.0%'
        elif fmt == "number":
            cell.number_format = '#,##0'


def write_formula_row(ws, row, label, formulas, fmt="currency", font=None, fill=None, indent=0):
    prefix = "  " * indent
    c = ws[f"A{row}"]
    c.value = prefix + label
    c.font = font or NORMAL_FONT
    if fill:
        for col in "ABCDEFG":
            ws[f"{col}{row}"].fill = fill
    for i, formula in enumerate(formulas):
        col = get_column_letter(3 + i)
        cell = ws[f"{col}{row}"]
        cell.value = formula
        cell.font = font or NORMAL_FONT
        cell.alignment = Alignment(horizontal="center")
        if fmt == "currency":
            cell.number_format = '#,##0'
        elif fmt == "percent":
            cell.number_format = '0.0%'
        elif fmt == "number":
            cell.number_format = '#,##0'
        if fill:
            cell.fill = fill


# =============================================================================
# SHEET: Combined P&L
# =============================================================================
ws = wb.active
ws.title = "P&L Model"
ws.sheet_properties.tabColor = "1A237E"

# Column widths
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 18
for col in "CDEFG":
    ws.column_dimensions[col].width = 16

r = 1

# ─── TITLE ────────────────────────────────────────────────────────────────────
style_title(ws, r, "VIA LABS + REALREAL GENUINE — P&L MODEL"); r += 1
style_title(ws, r, "All yellow cells are editable — change them and everything updates"); r += 2

# ─── ASSUMPTIONS: VIA LABS ────────────────────────────────────────────────────
ASSUMPTIONS_START = r
style_section_header(ws, r, "VIA LABS — ASSUMPTIONS"); r += 1
write_year_headers(ws, r); r += 1

# Row refs we'll need for formulas
R_MERCHANTS = r
write_editable_row(ws, r, "Active Merchants", [75, 350, 1500, 5000, 15000], "number", "Change these"); r += 1

r += 1  # spacer
ws[f"A{r}"].value = "Subscription Tier Pricing ($/month)"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_GROWTH_PRICE = r
write_editable_row(ws, r, "Growth tier price ($/mo)", [149, 149, 149, 149, 149], "currency"); r += 1
R_PRO_PRICE = r
write_editable_row(ws, r, "Professional tier price ($/mo)", [499, 499, 499, 499, 499], "currency"); r += 1
R_ENT_PRICE = r
write_editable_row(ws, r, "Enterprise tier price ($/mo)", [1499, 1499, 1499, 1499, 1499], "currency"); r += 1

r += 1
ws[f"A{r}"].value = "Subscription Mix (% of merchants on each tier)"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_FREE_MIX = r
write_editable_row(ws, r, "Starter (Free) %", [0.50, 0.40, 0.30, 0.25, 0.20], "percent"); r += 1
R_GROWTH_MIX = r
write_editable_row(ws, r, "Growth %", [0.30, 0.30, 0.30, 0.30, 0.30], "percent"); r += 1
R_PRO_MIX = r
write_editable_row(ws, r, "Professional %", [0.15, 0.20, 0.25, 0.25, 0.25], "percent"); r += 1
R_ENT_MIX = r
write_editable_row(ws, r, "Enterprise %", [0.05, 0.10, 0.15, 0.20, 0.25], "percent"); r += 1

r += 1
ws[f"A{r}"].value = "Micro Fee Rates"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_INTENT_FEE = r
write_editable_row(ws, r, "Intent fee ($ per intent)", [0.005, 0.005, 0.005, 0.005, 0.005], "currency")
for i in range(5):
    col = get_column_letter(3 + i)
    ws[f"{col}{r}"].number_format = '0.000'
r += 1

R_QUOTE_FEE = r
write_editable_row(ws, r, "Quote fee ($ per quote)", [0.01, 0.01, 0.01, 0.01, 0.01], "currency")
for i in range(5):
    col = get_column_letter(3 + i)
    ws[f"{col}{r}"].number_format = '0.000'
r += 1

R_TX_FEE_PCT = r
write_editable_row(ws, r, "Transaction fee (% of order)", [0.005, 0.005, 0.005, 0.005, 0.005], "percent"); r += 1

R_TX_FEE_CAP = r
write_editable_row(ws, r, "Transaction fee cap ($)", [2.00, 2.00, 2.00, 2.00, 2.00], "currency"); r += 1

r += 1
ws[f"A{r}"].value = "Transaction Volume Drivers"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_INTENTS_PER_M = r
write_editable_row(ws, r, "Intents per merchant per month", [200, 500, 800, 1200, 1800], "number"); r += 1
R_QUOTE_RATE = r
write_editable_row(ws, r, "Quote rate (% of intents)", [0.40, 0.45, 0.50, 0.55, 0.60], "percent"); r += 1
R_CONVERT_RATE = r
write_editable_row(ws, r, "Conversion rate (% of quotes)", [0.10, 0.12, 0.15, 0.18, 0.20], "percent"); r += 1
R_AOV = r
write_editable_row(ws, r, "Average order value ($)", [85, 95, 110, 120, 130], "currency"); r += 1

r += 1
ws[f"A{r}"].value = "Agent Top-Up"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_TOPUP_VOL = r
write_editable_row(ws, r, "Top-up volume ($)", [150000, 1200000, 8000000, 35000000, 120000000], "currency"); r += 1
R_TOPUP_MARGIN = r
write_editable_row(ws, r, "Top-up margin %", [0.025, 0.023, 0.020, 0.018, 0.015], "percent"); r += 1

r += 1
R_VIA_COGS_PCT = r
write_editable_row(ws, r, "Via Labs COGS (% of revenue)", [0.12, 0.12, 0.12, 0.12, 0.12], "percent", "Payment processing etc."); r += 1

r += 2

# ─── ASSUMPTIONS: RRG ─────────────────────────────────────────────────────────
style_section_header(ws, r, "REALREAL GENUINE — ASSUMPTIONS"); r += 1
write_year_headers(ws, r); r += 1

R_RRG_GMV = r
write_editable_row(ws, r, "Total GMV ($)", [180000, 950000, 4500000, 15000000, 45000000], "currency"); r += 1

r += 1
ws[f"A{r}"].value = "Commission Tiers (% of GMV in each tier)"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_T1_SHARE = r
write_editable_row(ws, r, "Tier 1 (under $100) — GMV share", [0.20, 0.15, 0.12, 0.10, 0.08], "percent"); r += 1
R_T1_RATE = r
write_editable_row(ws, r, "Tier 1 commission rate", [0.10, 0.10, 0.10, 0.10, 0.10], "percent"); r += 1

R_T2_SHARE = r
write_editable_row(ws, r, "Tier 2 ($100–$500) — GMV share", [0.40, 0.35, 0.33, 0.30, 0.27], "percent"); r += 1
R_T2_RATE = r
write_editable_row(ws, r, "Tier 2 commission rate", [0.07, 0.07, 0.07, 0.07, 0.07], "percent"); r += 1

R_T3_SHARE = r
write_editable_row(ws, r, "Tier 3 ($500–$2,000) — GMV share", [0.30, 0.35, 0.35, 0.35, 0.35], "percent"); r += 1
R_T3_RATE = r
write_editable_row(ws, r, "Tier 3 commission rate", [0.05, 0.05, 0.05, 0.05, 0.05], "percent"); r += 1

R_T4_SHARE = r
write_editable_row(ws, r, "Tier 4 ($2,000+) — GMV share", [0.10, 0.15, 0.20, 0.25, 0.30], "percent"); r += 1
R_T4_RATE = r
write_editable_row(ws, r, "Tier 4 commission rate", [0.03, 0.03, 0.03, 0.03, 0.03], "percent"); r += 1

r += 1
R_RRG_COCREATION = r
write_editable_row(ws, r, "Co-Creation Revenue ($)", [25000, 120000, 400000, 1000000, 2500000], "currency"); r += 1

R_RRG_COGS_PCT = r
write_editable_row(ws, r, "RRG COGS (% of revenue)", [0.18, 0.18, 0.18, 0.18, 0.18], "percent", "Fulfillment, payments etc."); r += 1

r += 2

# ─── ASSUMPTIONS: OPERATING EXPENSES ──────────────────────────────────────────
style_section_header(ws, r, "OPERATING EXPENSES — ASSUMPTIONS"); r += 1
write_year_headers(ws, r); r += 1

ws[f"A{r}"].value = "Via Labs OpEx"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_VIA_ENG = r
write_editable_row(ws, r, "Engineering & Product", [720000, 1440000, 2400000, 3600000, 5000000], "currency"); r += 1
R_VIA_OPS = r
write_editable_row(ws, r, "Operations & Admin", [180000, 360000, 600000, 900000, 1200000], "currency"); r += 1
R_VIA_MKT = r
write_editable_row(ws, r, "Sales & Marketing", [200000, 500000, 1000000, 2000000, 3500000], "currency"); r += 1
R_VIA_INFRA = r
write_editable_row(ws, r, "Infrastructure & Cloud", [96000, 200000, 400000, 750000, 1200000], "currency"); r += 1
R_VIA_LEGAL = r
write_editable_row(ws, r, "Legal & Compliance", [120000, 180000, 250000, 350000, 500000], "currency"); r += 1
R_VIA_GA = r
write_editable_row(ws, r, "General & Admin", [84000, 120000, 200000, 300000, 400000], "currency"); r += 1

r += 1
ws[f"A{r}"].value = "RealReal Genuine OpEx"
ws[f"A{r}"].font = HEADER_FONT; r += 1

R_RRG_DESIGN = r
write_editable_row(ws, r, "Design & Creative", [120000, 200000, 350000, 500000, 700000], "currency"); r += 1
R_RRG_BRAND = r
write_editable_row(ws, r, "Brand Partnerships", [60000, 120000, 250000, 400000, 600000], "currency"); r += 1
R_RRG_MKT = r
write_editable_row(ws, r, "Marketing & Community", [80000, 180000, 400000, 800000, 1500000], "currency"); r += 1
R_RRG_FULFILL = r
write_editable_row(ws, r, "Operations & Fulfillment", [100000, 200000, 450000, 900000, 1800000], "currency"); r += 1
R_RRG_TECH = r
write_editable_row(ws, r, "Platform & Tech", [40000, 80000, 150000, 250000, 400000], "currency"); r += 1
R_RRG_GA = r
write_editable_row(ws, r, "General & Admin", [50000, 70000, 100000, 150000, 250000], "currency"); r += 1

r += 1
R_RAISE = r
write_editable_row(ws, r, "Seed Raise Amount ($)", [5000000, None, None, None, None], "currency", "Starting cash"); r += 1

r += 2

# =============================================================================
# VIA LABS P&L (CALCULATED)
# =============================================================================
VIA_PL_START = r
style_section_header(ws, r, "VIA LABS — P&L (auto-calculated from assumptions above)"); r += 1
write_year_headers(ws, r); r += 1

# Subscription Revenue Lines
R_VIA_SUB_GROWTH = r
formulas = [f"={YEAR_COLS[yr]}{R_MERCHANTS}*{YEAR_COLS[yr]}{R_GROWTH_MIX}*{YEAR_COLS[yr]}{R_GROWTH_PRICE}*12" for yr in YEARS]
write_formula_row(ws, r, "Growth Tier Revenue", formulas, indent=1); r += 1

R_VIA_SUB_PRO = r
formulas = [f"={YEAR_COLS[yr]}{R_MERCHANTS}*{YEAR_COLS[yr]}{R_PRO_MIX}*{YEAR_COLS[yr]}{R_PRO_PRICE}*12" for yr in YEARS]
write_formula_row(ws, r, "Professional Tier Revenue", formulas, indent=1); r += 1

R_VIA_SUB_ENT = r
formulas = [f"={YEAR_COLS[yr]}{R_MERCHANTS}*{YEAR_COLS[yr]}{R_ENT_MIX}*{YEAR_COLS[yr]}{R_ENT_PRICE}*12" for yr in YEARS]
write_formula_row(ws, r, "Enterprise Tier Revenue", formulas, indent=1); r += 1

R_VIA_SUB_TOTAL = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_SUB_GROWTH}+{YEAR_COLS[yr]}{R_VIA_SUB_PRO}+{YEAR_COLS[yr]}{R_VIA_SUB_ENT}" for yr in YEARS]
write_formula_row(ws, r, "Total Subscription Revenue", formulas, font=RESULT_FONT, fill=TOTAL_FILL); r += 1

r += 1  # spacer

# Micro Fee calculated rows
R_VIA_TOTAL_INTENTS = r
formulas = [f"={YEAR_COLS[yr]}{R_MERCHANTS}*{YEAR_COLS[yr]}{R_INTENTS_PER_M}*12" for yr in YEARS]
write_formula_row(ws, r, "Total Intents (annual)", formulas, fmt="number", indent=1); r += 1

R_VIA_TOTAL_QUOTES = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_INTENTS}*{YEAR_COLS[yr]}{R_QUOTE_RATE}" for yr in YEARS]
write_formula_row(ws, r, "Total Quotes", formulas, fmt="number", indent=1); r += 1

R_VIA_TOTAL_TXS = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_QUOTES}*{YEAR_COLS[yr]}{R_CONVERT_RATE}" for yr in YEARS]
write_formula_row(ws, r, "Total Transactions", formulas, fmt="number", indent=1); r += 1

R_VIA_GMV = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_TXS}*{YEAR_COLS[yr]}{R_AOV}" for yr in YEARS]
write_formula_row(ws, r, "Platform GMV", formulas, indent=1); r += 1

r += 1

R_VIA_INTENT_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_INTENTS}*{YEAR_COLS[yr]}{R_INTENT_FEE}" for yr in YEARS]
write_formula_row(ws, r, "Intent Fee Revenue", formulas, indent=1); r += 1

R_VIA_QUOTE_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_QUOTES}*{YEAR_COLS[yr]}{R_QUOTE_FEE}" for yr in YEARS]
write_formula_row(ws, r, "Quote Fee Revenue", formulas, indent=1); r += 1

R_VIA_TX_REV = r
formulas = [f"=MIN({YEAR_COLS[yr]}{R_AOV}*{YEAR_COLS[yr]}{R_TX_FEE_PCT},{YEAR_COLS[yr]}{R_TX_FEE_CAP})*{YEAR_COLS[yr]}{R_VIA_TOTAL_TXS}" for yr in YEARS]
write_formula_row(ws, r, "Transaction Fee Revenue", formulas, indent=1); r += 1

R_VIA_MICRO_TOTAL = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_INTENT_REV}+{YEAR_COLS[yr]}{R_VIA_QUOTE_REV}+{YEAR_COLS[yr]}{R_VIA_TX_REV}" for yr in YEARS]
write_formula_row(ws, r, "Total Micro Fee Revenue", formulas, font=RESULT_FONT, fill=TOTAL_FILL); r += 1

r += 1

R_VIA_TOPUP_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_TOPUP_VOL}*{YEAR_COLS[yr]}{R_TOPUP_MARGIN}" for yr in YEARS]
write_formula_row(ws, r, "Agent Top-Up Margin Revenue", formulas, font=RESULT_FONT, fill=TOTAL_FILL); r += 1

r += 1

R_VIA_TOTAL_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_SUB_TOTAL}+{YEAR_COLS[yr]}{R_VIA_MICRO_TOTAL}+{YEAR_COLS[yr]}{R_VIA_TOPUP_REV}" for yr in YEARS]
write_formula_row(ws, r, "VIA LABS TOTAL REVENUE", formulas, font=Font(name="Arial", size=11, bold=True), fill=TOTAL_FILL); r += 1

R_VIA_COGS = r
formulas = [f"=-{YEAR_COLS[yr]}{R_VIA_TOTAL_REV}*{YEAR_COLS[yr]}{R_VIA_COGS_PCT}" for yr in YEARS]
write_formula_row(ws, r, "Cost of Revenue", formulas); r += 1

R_VIA_GP = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_REV}+{YEAR_COLS[yr]}{R_VIA_COGS}" for yr in YEARS]
write_formula_row(ws, r, "Gross Profit", formulas, font=RESULT_FONT); r += 1

R_VIA_GM = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_VIA_TOTAL_REV}=0,0,{YEAR_COLS[yr]}{R_VIA_GP}/{YEAR_COLS[yr]}{R_VIA_TOTAL_REV})" for yr in YEARS]
write_formula_row(ws, r, "Gross Margin %", formulas, fmt="percent"); r += 1

r += 1

R_VIA_TOTAL_OPEX = r
formulas = [f"=-({YEAR_COLS[yr]}{R_VIA_ENG}+{YEAR_COLS[yr]}{R_VIA_OPS}+{YEAR_COLS[yr]}{R_VIA_MKT}+{YEAR_COLS[yr]}{R_VIA_INFRA}+{YEAR_COLS[yr]}{R_VIA_LEGAL}+{YEAR_COLS[yr]}{R_VIA_GA})" for yr in YEARS]
write_formula_row(ws, r, "Total Operating Expenses", formulas, font=RESULT_FONT); r += 1

R_VIA_EBITDA = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_GP}+{YEAR_COLS[yr]}{R_VIA_TOTAL_OPEX}" for yr in YEARS]
write_formula_row(ws, r, "VIA LABS EBITDA", formulas, font=Font(name="Arial", size=11, bold=True), fill=TOTAL_FILL); r += 1

R_VIA_EBITDA_M = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_VIA_TOTAL_REV}=0,0,{YEAR_COLS[yr]}{R_VIA_EBITDA}/{YEAR_COLS[yr]}{R_VIA_TOTAL_REV})" for yr in YEARS]
write_formula_row(ws, r, "EBITDA Margin %", formulas, fmt="percent"); r += 1

r += 2

# =============================================================================
# RRG P&L (CALCULATED)
# =============================================================================
style_section_header(ws, r, "REALREAL GENUINE — P&L (auto-calculated from assumptions above)"); r += 1
write_year_headers(ws, r); r += 1

R_RRG_T1_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_GMV}*{YEAR_COLS[yr]}{R_T1_SHARE}*{YEAR_COLS[yr]}{R_T1_RATE}" for yr in YEARS]
write_formula_row(ws, r, "Tier 1 Commission (under $100)", formulas, indent=1); r += 1

R_RRG_T2_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_GMV}*{YEAR_COLS[yr]}{R_T2_SHARE}*{YEAR_COLS[yr]}{R_T2_RATE}" for yr in YEARS]
write_formula_row(ws, r, "Tier 2 Commission ($100–$500)", formulas, indent=1); r += 1

R_RRG_T3_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_GMV}*{YEAR_COLS[yr]}{R_T3_SHARE}*{YEAR_COLS[yr]}{R_T3_RATE}" for yr in YEARS]
write_formula_row(ws, r, "Tier 3 Commission ($500–$2K)", formulas, indent=1); r += 1

R_RRG_T4_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_GMV}*{YEAR_COLS[yr]}{R_T4_SHARE}*{YEAR_COLS[yr]}{R_T4_RATE}" for yr in YEARS]
write_formula_row(ws, r, "Tier 4 Commission ($2K+)", formulas, indent=1); r += 1

R_RRG_COMM_TOTAL = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_T1_REV}+{YEAR_COLS[yr]}{R_RRG_T2_REV}+{YEAR_COLS[yr]}{R_RRG_T3_REV}+{YEAR_COLS[yr]}{R_RRG_T4_REV}" for yr in YEARS]
write_formula_row(ws, r, "Total Platform Commission", formulas, font=RESULT_FONT, fill=TOTAL_FILL); r += 1

R_RRG_BLEND = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_RRG_GMV}=0,0,{YEAR_COLS[yr]}{R_RRG_COMM_TOTAL}/{YEAR_COLS[yr]}{R_RRG_GMV})" for yr in YEARS]
write_formula_row(ws, r, "Blended Commission Rate", formulas, fmt="percent"); r += 1

r += 1

R_RRG_COCREATION_ROW = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_COCREATION}" for yr in YEARS]
write_formula_row(ws, r, "Co-Creation Revenue", formulas, font=RESULT_FONT, fill=TOTAL_FILL); r += 1

r += 1

R_RRG_TOTAL_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_COMM_TOTAL}+{YEAR_COLS[yr]}{R_RRG_COCREATION_ROW}" for yr in YEARS]
write_formula_row(ws, r, "RRG TOTAL REVENUE", formulas, font=Font(name="Arial", size=11, bold=True), fill=TOTAL_FILL); r += 1

R_RRG_COGS = r
formulas = [f"=-{YEAR_COLS[yr]}{R_RRG_TOTAL_REV}*{YEAR_COLS[yr]}{R_RRG_COGS_PCT}" for yr in YEARS]
write_formula_row(ws, r, "Cost of Revenue", formulas); r += 1

R_RRG_GP = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_TOTAL_REV}+{YEAR_COLS[yr]}{R_RRG_COGS}" for yr in YEARS]
write_formula_row(ws, r, "Gross Profit", formulas, font=RESULT_FONT); r += 1

R_RRG_GM = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_RRG_TOTAL_REV}=0,0,{YEAR_COLS[yr]}{R_RRG_GP}/{YEAR_COLS[yr]}{R_RRG_TOTAL_REV})" for yr in YEARS]
write_formula_row(ws, r, "Gross Margin %", formulas, fmt="percent"); r += 1

r += 1

R_RRG_TOTAL_OPEX = r
formulas = [f"=-({YEAR_COLS[yr]}{R_RRG_DESIGN}+{YEAR_COLS[yr]}{R_RRG_BRAND}+{YEAR_COLS[yr]}{R_RRG_MKT}+{YEAR_COLS[yr]}{R_RRG_FULFILL}+{YEAR_COLS[yr]}{R_RRG_TECH}+{YEAR_COLS[yr]}{R_RRG_GA})" for yr in YEARS]
write_formula_row(ws, r, "Total Operating Expenses", formulas, font=RESULT_FONT); r += 1

R_RRG_EBITDA = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_GP}+{YEAR_COLS[yr]}{R_RRG_TOTAL_OPEX}" for yr in YEARS]
write_formula_row(ws, r, "RRG EBITDA", formulas, font=Font(name="Arial", size=11, bold=True), fill=TOTAL_FILL); r += 1

R_RRG_EBITDA_M = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_RRG_TOTAL_REV}=0,0,{YEAR_COLS[yr]}{R_RRG_EBITDA}/{YEAR_COLS[yr]}{R_RRG_TOTAL_REV})" for yr in YEARS]
write_formula_row(ws, r, "EBITDA Margin %", formulas, fmt="percent"); r += 1

r += 2

# =============================================================================
# COMBINED P&L
# =============================================================================
style_section_header(ws, r, "COMBINED GROUP P&L"); r += 1
write_year_headers(ws, r); r += 1

R_COMB_VIA_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_REV}" for yr in YEARS]
write_formula_row(ws, r, "Via Labs Revenue", formulas, indent=1); r += 1

R_COMB_RRG_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_RRG_TOTAL_REV}" for yr in YEARS]
write_formula_row(ws, r, "RealReal Genuine Revenue", formulas, indent=1); r += 1

R_COMB_REV = r
formulas = [f"={YEAR_COLS[yr]}{R_COMB_VIA_REV}+{YEAR_COLS[yr]}{R_COMB_RRG_REV}" for yr in YEARS]
write_formula_row(ws, r, "COMBINED REVENUE", formulas, font=Font(name="Arial", size=12, bold=True), fill=TOTAL_FILL); r += 1

r += 1

# Revenue mix
R_VIA_MIX = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_COMB_REV}=0,0,{YEAR_COLS[yr]}{R_COMB_VIA_REV}/{YEAR_COLS[yr]}{R_COMB_REV})" for yr in YEARS]
write_formula_row(ws, r, "Via Labs % of revenue", formulas, fmt="percent"); r += 1

R_RRG_MIX_ROW = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_COMB_REV}=0,0,{YEAR_COLS[yr]}{R_COMB_RRG_REV}/{YEAR_COLS[yr]}{R_COMB_REV})" for yr in YEARS]
write_formula_row(ws, r, "RRG % of revenue", formulas, fmt="percent"); r += 1

r += 1

R_COMB_COGS = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_COGS}+{YEAR_COLS[yr]}{R_RRG_COGS}" for yr in YEARS]
write_formula_row(ws, r, "Combined Cost of Revenue", formulas); r += 1

R_COMB_GP = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_GP}+{YEAR_COLS[yr]}{R_RRG_GP}" for yr in YEARS]
write_formula_row(ws, r, "Combined Gross Profit", formulas, font=RESULT_FONT); r += 1

R_COMB_GM = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_COMB_REV}=0,0,{YEAR_COLS[yr]}{R_COMB_GP}/{YEAR_COLS[yr]}{R_COMB_REV})" for yr in YEARS]
write_formula_row(ws, r, "Gross Margin %", formulas, fmt="percent"); r += 1

r += 1

R_COMB_OPEX = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_TOTAL_OPEX}+{YEAR_COLS[yr]}{R_RRG_TOTAL_OPEX}" for yr in YEARS]
write_formula_row(ws, r, "Combined Operating Expenses", formulas, font=RESULT_FONT); r += 1

r += 1

R_COMB_EBITDA = r
formulas = [f"={YEAR_COLS[yr]}{R_VIA_EBITDA}+{YEAR_COLS[yr]}{R_RRG_EBITDA}" for yr in YEARS]
write_formula_row(ws, r, "COMBINED EBITDA", formulas, font=Font(name="Arial", size=12, bold=True), fill=TOTAL_FILL); r += 1

R_COMB_EBITDA_M = r
formulas = [f"=IF({YEAR_COLS[yr]}{R_COMB_REV}=0,0,{YEAR_COLS[yr]}{R_COMB_EBITDA}/{YEAR_COLS[yr]}{R_COMB_REV})" for yr in YEARS]
write_formula_row(ws, r, "EBITDA Margin %", formulas, fmt="percent"); r += 1

r += 2

# Cash position
style_section_header(ws, r, "CASH POSITION"); r += 1
write_year_headers(ws, r); r += 1

R_CASH = r
# Year 1: Raise + EBITDA. Year 2+: Prior year cash + EBITDA
formulas = []
for i, yr in enumerate(YEARS):
    col = YEAR_COLS[yr]
    if i == 0:
        formulas.append(f"=C{R_RAISE}+{col}{R_COMB_EBITDA}")
    else:
        prev_col = YEAR_COLS[YEARS[i - 1]]
        formulas.append(f"={prev_col}{R_CASH}+{col}{R_COMB_EBITDA}")
write_formula_row(ws, r, "Cash Balance", formulas, font=Font(name="Arial", size=11, bold=True), fill=TOTAL_FILL); r += 1

r += 1
R_ARR_MULT = r
formulas = [f"=IF(C{R_RAISE}=0,0,{YEAR_COLS[yr]}{R_COMB_REV}/C{R_RAISE})" for yr in YEARS]
write_formula_row(ws, r, "ARR / Raise Multiple", formulas, fmt="number")
for i in range(5):
    col = get_column_letter(3 + i)
    ws[f"{col}{r}"].number_format = '0.0"x"'
r += 1


# ─── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = "C1"

# ─── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Done: {OUTPUT}")
print("Upload this file to Google Drive → Open with Google Sheets")
print("All yellow cells are editable — change any assumption and the P&L updates automatically.")
